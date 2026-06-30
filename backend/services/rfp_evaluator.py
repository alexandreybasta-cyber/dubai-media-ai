"""
RFP Evaluator Service - AI-powered vendor response evaluation using Qwen via DashScope.
"""

import asyncio
import io
import json
import logging
from typing import Any

import httpx
import pdfplumber
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from config import settings

logger = logging.getLogger(__name__)

EVAL_SYSTEM_PROMPT = """You are an expert RFP evaluator for Dubai Media Incorporated. 
You analyze vendor proposals against specific evaluation criteria with precision and objectivity.
You provide fair, evidence-based scoring with clear justifications.
Always output valid JSON as requested."""


class RFPEvaluator:
    """AI-powered RFP response evaluator using Qwen via DashScope."""

    def __init__(self):
        self.api_url = f"{settings.DASHSCOPE_BASE_URL}/chat/completions"
        self.model = settings.MODEL_TEXT
        self.api_key = settings.DASHSCOPE_API_KEY
        self.max_retries = 3

    async def _call_llm(self, messages: list[dict], temperature: float = 0.3) -> str:
        """Call DashScope API with retries and exponential backoff."""
        if not self.api_key:
            raise ValueError(
                "DASHSCOPE_API_KEY is not configured. Please set it in your .env file."
            )

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": 4096,
        }

        last_error = None
        for attempt in range(self.max_retries):
            try:
                async with httpx.AsyncClient(timeout=120.0) as client:
                    response = await client.post(
                        self.api_url, headers=headers, json=payload
                    )

                    if response.status_code == 429:
                        wait_time = (2**attempt) * 2
                        logger.warning(
                            f"Rate limited, retrying in {wait_time}s (attempt {attempt + 1})"
                        )
                        await asyncio.sleep(wait_time)
                        continue

                    if response.status_code != 200:
                        error_text = response.text
                        logger.error(f"DashScope API error {response.status_code}: {error_text}")
                        last_error = RuntimeError(
                            f"DashScope API returned {response.status_code}: {error_text}"
                        )
                        await asyncio.sleep(2**attempt)
                        continue

                    data = response.json()
                    content = data["choices"][0]["message"]["content"]
                    return content

            except httpx.TimeoutException:
                last_error = RuntimeError("DashScope API request timed out")
                logger.warning(f"Timeout on attempt {attempt + 1}")
                await asyncio.sleep(2**attempt)
            except Exception as e:
                last_error = RuntimeError(f"DashScope API call failed: {str(e)}")
                logger.error(f"Error on attempt {attempt + 1}: {e}")
                await asyncio.sleep(2**attempt)

        raise last_error or RuntimeError("All API retries exhausted")

    def extract_text_from_pdf(self, pdf_bytes: bytes) -> str:
        """Extract text from PDF using pdfplumber."""
        text_parts = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)

    def extract_text_from_docx(self, docx_bytes: bytes) -> str:
        """Extract text from DOCX using python-docx."""
        doc = Document(io.BytesIO(docx_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def extract_text(self, file_bytes: bytes, filename: str) -> str:
        """Extract text from a file based on its extension."""
        lower = filename.lower()
        if lower.endswith(".pdf"):
            return self.extract_text_from_pdf(file_bytes)
        elif lower.endswith(".docx"):
            return self.extract_text_from_docx(file_bytes)
        else:
            # Try to decode as plain text
            return file_bytes.decode("utf-8", errors="replace")

    async def evaluate_single_vendor(
        self, rfp_text: str, vendor_name: str, response_text: str, criteria: list[dict]
    ) -> dict:
        """Evaluate a single vendor's response against the criteria."""
        criteria_json = json.dumps(criteria, indent=2)

        # Truncate texts if too long to fit in context
        max_rfp_len = 6000
        max_response_len = 8000
        rfp_excerpt = rfp_text[:max_rfp_len] + ("..." if len(rfp_text) > max_rfp_len else "")
        response_excerpt = response_text[:max_response_len] + (
            "..." if len(response_text) > max_response_len else ""
        )

        user_prompt = f"""Evaluate this vendor's response against the RFP criteria.

ORIGINAL RFP:
{rfp_excerpt}

VENDOR RESPONSE ({vendor_name}):
{response_excerpt}

EVALUATION CRITERIA:
{criteria_json}

Score each criterion from 1-10 and provide detailed analysis. Output ONLY valid JSON in this exact format:
{{
  "scores": [
    {{
      "criterion": "<criterion name>",
      "score": <1-10>,
      "justification": "<2-3 sentences explaining the score>",
      "evidence": "<direct quote or reference from the vendor response>"
    }}
  ],
  "strengths": ["<strength 1>", "<strength 2>", "<strength 3>"],
  "gaps": ["<gap 1>", "<gap 2>"],
  "risks": ["<risk 1>", "<risk 2>"],
  "mandatory_compliance": [
    {{
      "requirement": "<mandatory requirement if any>",
      "status": "pass",
      "note": "<brief note>"
    }}
  ]
}}"""

        messages = [
            {"role": "system", "content": EVAL_SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ]

        raw_response = await self._call_llm(messages)

        # Parse JSON from response (handle markdown code blocks)
        json_str = raw_response.strip()
        if json_str.startswith("```"):
            lines = json_str.split("\n")
            # Remove first and last lines (```json and ```)
            lines = [l for l in lines if not l.strip().startswith("```")]
            json_str = "\n".join(lines)

        try:
            result = json.loads(json_str)
        except json.JSONDecodeError:
            # Try to find JSON in the response
            start = json_str.find("{")
            end = json_str.rfind("}") + 1
            if start >= 0 and end > start:
                try:
                    result = json.loads(json_str[start:end])
                except json.JSONDecodeError:
                    logger.error(f"Failed to parse evaluation JSON for {vendor_name}")
                    result = self._fallback_evaluation(criteria)
            else:
                result = self._fallback_evaluation(criteria)

        return result

    def _fallback_evaluation(self, criteria: list[dict]) -> dict:
        """Generate a fallback evaluation when AI parsing fails."""
        return {
            "scores": [
                {
                    "criterion": c["name"],
                    "score": 5,
                    "justification": "Unable to evaluate automatically. Manual review recommended.",
                    "evidence": "N/A",
                }
                for c in criteria
            ],
            "strengths": ["Requires manual review"],
            "gaps": ["AI evaluation incomplete"],
            "risks": ["Manual verification needed"],
            "mandatory_compliance": [],
        }

    async def evaluate_responses(
        self, rfp_text: str, vendor_responses: list[dict], criteria: list[dict]
    ) -> dict:
        """
        Evaluate all vendor responses against RFP criteria.

        Args:
            rfp_text: The original RFP text content
            vendor_responses: List of {vendor_name: str, response_text: str}
            criteria: List of {name: str, weight: float, description: str, mandatory: bool}

        Returns evaluation results dict.
        """
        vendors_results = []

        for vendor in vendor_responses:
            vendor_name = vendor["vendor_name"]
            response_text = vendor["response_text"]

            eval_result = await self.evaluate_single_vendor(
                rfp_text, vendor_name, response_text, criteria
            )

            # Calculate weighted total
            total_weight = sum(c.get("weight", 0) for c in criteria)
            weighted_total = 0.0
            if total_weight > 0:
                for score_item in eval_result.get("scores", []):
                    # Match score to criterion weight
                    criterion_name = score_item.get("criterion", "")
                    weight = next(
                        (
                            c["weight"]
                            for c in criteria
                            if c["name"].lower() == criterion_name.lower()
                        ),
                        0,
                    )
                    weighted_total += (score_item.get("score", 0) * weight) / total_weight

            # Normalize to 0-100 scale
            weighted_total = round(weighted_total * 10, 1)

            vendors_results.append(
                {
                    "vendor_name": vendor_name,
                    "scores": eval_result.get("scores", []),
                    "weighted_total": weighted_total,
                    "strengths": eval_result.get("strengths", []),
                    "gaps": eval_result.get("gaps", []),
                    "risks": eval_result.get("risks", []),
                    "mandatory_compliance": eval_result.get("mandatory_compliance", []),
                }
            )

        # Generate overall recommendation
        recommendation = await self._generate_recommendation(vendors_results, criteria)

        # Generate follow-up questions
        follow_up = await self._generate_follow_up_questions(vendors_results)

        return {
            "vendors": vendors_results,
            "recommendation": recommendation,
            "follow_up_questions": follow_up,
        }

    async def _generate_recommendation(
        self, vendors_results: list[dict], criteria: list[dict]
    ) -> str:
        """Generate an AI narrative recommendation."""
        summary = []
        for v in vendors_results:
            summary.append(f"- {v['vendor_name']}: Weighted Score {v['weighted_total']}/100")

        prompt = f"""Based on these evaluation results, provide a professional recommendation narrative (3-5 paragraphs):

Vendor Scores:
{chr(10).join(summary)}

Criteria evaluated: {', '.join(c['name'] for c in criteria)}

Write a clear, professional recommendation that:
1. States the recommended vendor and why
2. Summarizes key differentiators
3. Notes important risks or considerations
4. Suggests next steps

Output plain text (not JSON)."""

        messages = [
            {"role": "system", "content": EVAL_SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ]

        try:
            return await self._call_llm(messages, temperature=0.5)
        except Exception as e:
            logger.error(f"Recommendation generation failed: {e}")
            # Sort by weighted total
            sorted_vendors = sorted(vendors_results, key=lambda x: x["weighted_total"], reverse=True)
            if sorted_vendors:
                return f"Based on the evaluation scores, {sorted_vendors[0]['vendor_name']} achieved the highest weighted score of {sorted_vendors[0]['weighted_total']}/100 and is recommended for further consideration."
            return "Unable to generate automated recommendation. Please review the individual scores."

    async def _generate_follow_up_questions(self, vendors_results: list[dict]) -> dict:
        """Generate follow-up questions for each vendor."""
        follow_up = {}
        for v in vendors_results:
            gaps = v.get("gaps", [])
            risks = v.get("risks", [])
            questions = []
            for gap in gaps[:3]:
                questions.append(f"Can you provide more detail on: {gap}?")
            for risk in risks[:2]:
                questions.append(f"How do you plan to mitigate: {risk}?")
            if not questions:
                questions = ["Please provide references from similar projects."]
            follow_up[v["vendor_name"]] = questions
        return follow_up

    def export_xlsx(self, evaluation_results: dict) -> bytes:
        """Generate comparison matrix XLSX."""
        wb = Workbook()

        # ─── Sheet 1: Comparison Matrix ─────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Comparison Matrix"

        vendors = evaluation_results.get("vendors", [])
        if not vendors:
            ws1.append(["No evaluation data available"])
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Header row
        header = ["Criterion", "Weight (%)"] + [v["vendor_name"] for v in vendors]
        ws1.append(header)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        for col in range(1, len(header) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Score color fills
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        # Data rows - use first vendor's scores as criteria reference
        first_vendor_scores = vendors[0].get("scores", [])
        for i, score_item in enumerate(first_vendor_scores):
            criterion = score_item.get("criterion", f"Criterion {i+1}")
            # Find weight from scores
            row = [criterion, ""]
            for v in vendors:
                v_scores = v.get("scores", [])
                vendor_score = next(
                    (s["score"] for s in v_scores if s.get("criterion") == criterion),
                    0,
                )
                row.append(vendor_score)
            ws1.append(row)

            # Apply color formatting to score cells
            row_num = ws1.max_row
            for col in range(3, len(header) + 1):
                cell = ws1.cell(row=row_num, column=col)
                score_val = cell.value or 0
                if score_val >= 8:
                    cell.fill = green_fill
                elif score_val >= 5:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
                cell.alignment = Alignment(horizontal="center")

        # Weighted Total row
        total_row = ["WEIGHTED TOTAL", "100%"] + [
            v.get("weighted_total", 0) for v in vendors
        ]
        ws1.append(total_row)
        total_row_num = ws1.max_row
        for col in range(1, len(header) + 1):
            cell = ws1.cell(row=total_row_num, column=col)
            cell.font = Font(bold=True, size=12)

        # Column widths
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 12
        for col_idx in range(3, len(header) + 1):
            ws1.column_dimensions[chr(64 + col_idx)].width = 18

        # ─── Sheet 2: Detailed Scores ───────────────────────────────────────
        ws2 = wb.create_sheet("Detailed Scores")
        for v in vendors:
            ws2.append([v["vendor_name"]])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14)
            ws2.append(["Criterion", "Score", "Justification"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
            ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)
            ws2.cell(row=ws2.max_row, column=3).font = Font(bold=True)

            for s in v.get("scores", []):
                ws2.append([
                    s.get("criterion", ""),
                    s.get("score", 0),
                    s.get("justification", ""),
                ])
            ws2.append([])  # Blank row separator

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 60

        # ─── Sheet 3: Recommendation ────────────────────────────────────────
        ws3 = wb.create_sheet("Recommendation")
        ws3.append(["AI Recommendation"])
        ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws3.append([])
        ws3.append([evaluation_results.get("recommendation", "N/A")])
        ws3.append([])
        ws3.append(["Follow-up Questions"])
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12)

        follow_up = evaluation_results.get("follow_up_questions", {})
        for vendor_name, questions in follow_up.items():
            ws3.append([])
            ws3.append([vendor_name])
            ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
            for q in questions:
                ws3.append([f"  • {q}"])

        ws3.column_dimensions["A"].width = 80

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_pdf_report(self, evaluation_results: dict) -> bytes:
        """Generate comprehensive PDF evaluation report."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=cm * 2,
            leftMargin=cm * 2,
            topMargin=cm * 2,
            bottomMargin=cm * 2,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor("#F97316"),
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor("#1F2937"),
        )
        body_style = styles["Normal"]
        body_style.fontSize = 10
        body_style.leading = 14

        # ─── Cover Page ─────────────────────────────────────────────────────
        elements.append(Spacer(1, cm * 4))
        elements.append(Paragraph("RFP Evaluation Report", title_style))
        elements.append(Spacer(1, cm * 1))
        elements.append(Paragraph("Dubai Media Incorporated", heading_style))
        elements.append(Spacer(1, cm * 1))

        vendors = evaluation_results.get("vendors", [])
        vendor_names = ", ".join(v["vendor_name"] for v in vendors)
        elements.append(Paragraph(f"Vendors Evaluated: {vendor_names}", body_style))
        elements.append(Paragraph(f"Total Vendors: {len(vendors)}", body_style))
        elements.append(PageBreak())

        # ─── Executive Summary ──────────────────────────────────────────────
        elements.append(Paragraph("Executive Summary", heading_style))
        recommendation = evaluation_results.get("recommendation", "No recommendation available.")
        # Split recommendation into paragraphs
        for para in recommendation.split("\n"):
            if para.strip():
                elements.append(Paragraph(para.strip(), body_style))
                elements.append(Spacer(1, 6))
        elements.append(Spacer(1, cm * 1))

        # ─── Comparison Table ───────────────────────────────────────────────
        elements.append(Paragraph("Score Comparison", heading_style))

        if vendors:
            # Build table data
            table_header = ["Criterion"] + [v["vendor_name"] for v in vendors]
            table_data = [table_header]

            first_scores = vendors[0].get("scores", [])
            for score_item in first_scores:
                criterion = score_item.get("criterion", "")
                row = [criterion]
                for v in vendors:
                    v_score = next(
                        (
                            s["score"]
                            for s in v.get("scores", [])
                            if s.get("criterion") == criterion
                        ),
                        0,
                    )
                    row.append(str(v_score))
                table_data.append(row)

            # Totals row
            total_row = ["WEIGHTED TOTAL"] + [
                str(v.get("weighted_total", 0)) for v in vendors
            ]
            table_data.append(total_row)

            col_widths = [cm * 5] + [cm * 3.5] * len(vendors)
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF7ED")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, cm * 1))

        # ─── Per-Vendor Scorecards ──────────────────────────────────────────
        for v in vendors:
            elements.append(PageBreak())
            elements.append(Paragraph(f"Vendor: {v['vendor_name']}", heading_style))
            elements.append(
                Paragraph(f"Weighted Score: {v.get('weighted_total', 0)}/100", body_style)
            )
            elements.append(Spacer(1, 12))

            # Strengths
            elements.append(Paragraph("Strengths:", ParagraphStyle("Bold", parent=body_style, fontName="Helvetica-Bold")))
            for s in v.get("strengths", []):
                elements.append(Paragraph(f"  • {s}", body_style))

            elements.append(Spacer(1, 8))

            # Gaps
            elements.append(Paragraph("Gaps:", ParagraphStyle("BoldGap", parent=body_style, fontName="Helvetica-Bold")))
            for g in v.get("gaps", []):
                elements.append(Paragraph(f"  • {g}", body_style))

            elements.append(Spacer(1, 8))

            # Risks
            elements.append(Paragraph("Risks:", ParagraphStyle("BoldRisk", parent=body_style, fontName="Helvetica-Bold")))
            for r in v.get("risks", []):
                elements.append(Paragraph(f"  • {r}", body_style))

        # ─── Footer ─────────────────────────────────────────────────────────
        elements.append(Spacer(1, cm * 2))
        footer_style = ParagraphStyle(
            "Footer", parent=body_style, fontSize=8, textColor=colors.grey
        )
        elements.append(
            Paragraph(
                "This evaluation was generated by Qwen AI. Results should be reviewed by procurement professionals.",
                footer_style,
            )
        )

        doc.build(elements)
        return buffer.getvalue()
